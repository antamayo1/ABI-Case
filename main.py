from openpyxl import load_workbook
import streamlit as st
from openai import OpenAI
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.cell.text import InlineFont
from openpyxl.drawing.image import Image
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from io import BytesIO
from functools import cache
from PIL import Image as PILImage
from collections import OrderedDict
import streamlit_sortables as sortables
from openpyxl.styles import Alignment, Font
from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont
from openpyxl.worksheet.pagebreak import Break
import pandas as pd

MAX_MODEL_CHARACTERS_PER_COLUMN = 30
MAX_SUPPLIER_CHARACTERS_PER_COLUMN = 24
AVAILABLE_HEIGHT_PER_PAGE = 720
if 'total_height' not in st.session_state:
  st.session_state.total_height = 0
AI_outputs = {}

AI_client = OpenAI(
  api_key=st.secrets["OPEN_AI_KEY"]
)

fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
sec_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

st.set_page_config(page_title="Design Transformer", layout="wide", page_icon="https://images.squarespace-cdn.com/content/v1/64de080086e544329d5ae8ad/4cc35b7d-6693-4a6e-a547-736c1ebbc63a/favicon.ico?format=100w")

def count_wrapped_lines(text, max_chars):
  words = text.split()
  lines = 1
  current_len = 0
  for word in words:
    add_len = len(word) + (1 if current_len > 0 else 0)
    if current_len + add_len > max_chars:
      lines += 1
      current_len = len(word)
    else:
      current_len += add_len
  return lines

def get_image_anchor(img):
  anchor = img.anchor
  if hasattr(anchor, '_from'):
    col = anchor._from.col + 1
    row = anchor._from.row + 1
    return f"{get_column_letter(col)}{row}"
  elif hasattr(anchor, 'cell'):
    return anchor.cell
  else:
    return str(anchor)

def create_blank_image():
  img = PILImage.new('RGBA', (1, 1), (255, 255, 255, 0))
  buf = BytesIO()
  img.save(buf, format='PNG')
  buf.seek(0)
  return Image(buf)

def pad_images_with_blanks(images, total_rows):
  expected_cells = [f'E{idx}' for idx in range(11, total_rows + 11)]
  actual_anchors = [get_image_anchor(img) for img in images]
  padded_images = []
  img_idx = 0
  for cell in expected_cells:
    if img_idx < len(images) and actual_anchors[img_idx] == cell:
      padded_images.append(images[img_idx])
      img_idx += 1
    else:
      padded_images.append(create_blank_image())
  return padded_images

def box(worksheet, start_cell, end_cell, border_style="thin"):
  side = Side(border_style=border_style, color="000000")
  start_row, start_col = coordinate_to_tuple(start_cell)
  end_row, end_col = coordinate_to_tuple(end_cell)
  min_row, max_row = min(start_row, end_row), max(start_row, end_row)
  min_col, max_col = min(start_col, end_col), max(start_col, end_col)
  for row in range(min_row, max_row + 1):
    for col in range(min_col, max_col + 1):
      cell = worksheet.cell(row=row, column=col)
      cell_border = Border(
        left=side if col == min_col else None,
        right=side if col == max_col else None,
        top=side if row == min_row else None,
        bottom=side if row == max_row else None,
      )
      cell.border = Border(
        left=cell_border.left or cell.border.left,
        right=cell_border.right or cell.border.right,
        top=cell_border.top or cell.border.top,
        bottom=cell_border.bottom or cell.border.bottom,
      )
  return worksheet

def box_fill(worksheet, start_cell, end_cell, border_style="thin"):
  side = Side(border_style=border_style, color="000000")
  border = Border(left=side, right=side, top=side, bottom=side)

  start_row, start_col = coordinate_to_tuple(start_cell)
  end_row, end_col = coordinate_to_tuple(end_cell)
  min_row, max_row = min(start_row, end_row), max(start_row, end_row)
  min_col, max_col = min(start_col, end_col), max(start_col, end_col)

  for row in range(min_row, max_row + 1):
    for col in range(min_col, max_col + 1):
      cell = worksheet.cell(row=row, column=col)
      cell.border = border
  return worksheet

def padLogoImage(img):
  pil_img = PILImage.open(img)
  padding_left = 160
  padding_right = 160
  padding_top = 120
  padding_bottom = 120
  new_width = pil_img.width + padding_left + padding_right
  new_height = pil_img.height + padding_top + padding_bottom
  padded_img = PILImage.new('RGBA', (new_width, new_height), (255, 255, 255, 0))
  padded_img.paste(pil_img, (padding_left, padding_top), pil_img if pil_img.mode == 'RGBA' else None)
  buf = BytesIO()
  padded_img.save(buf, format='PNG')
  buf.seek(0)
  img = Image(buf)
  img.width = 195
  img.height = 95
  return img

def padProductImage(image, row_height):
  pil_img = PILImage.open(image.ref)
  
  padding_left = 5
  padding_right = 5
  padding_top = 5
  padding_bottom = 5
  
  new_width = pil_img.width + padding_left + padding_right
  new_height = pil_img.height + padding_top + padding_bottom
  padded_img = PILImage.new('RGBA', (new_width, new_height), (255, 255, 255, 0))
  
  padded_img.paste(pil_img, (padding_left, padding_top), pil_img if pil_img.mode == 'RGBA' else None)
  
  buf = BytesIO()
  padded_img.save(buf, format='PNG')
  buf.seek(0)
  
  new_img = Image(buf)

  new_img.width = 182
  new_img.height = row_height * 1.325
  
  return new_img

def columnRowDimensions(worksheet):
  worksheet.column_dimensions['A'].width = 14.78 * 1.4
  worksheet.column_dimensions['B'].width = 4.56 * 1.4
  worksheet.column_dimensions['C'].width = 20.56 * 1.4
  worksheet.column_dimensions['D'].width = 6.89 * 1.4
  worksheet.column_dimensions['E'].width = 18.11 * 1.4
  worksheet.column_dimensions['F'].width = 25.33 * 1.4
  worksheet.column_dimensions['G'].width = 19.89 * 1.4
  worksheet.column_dimensions['H'].width = 10.11 * 1.4
  worksheet.column_dimensions['I'].width = 9.56 * 1.4
  worksheet.column_dimensions['J'].width = 11.11 * 1.4
  worksheet.column_dimensions['K'].width = 40.11 * 1.4
  worksheet.row_dimensions[1].height = 18
  worksheet.row_dimensions[2].height = 18
  worksheet.row_dimensions[3].height = 18
  worksheet.row_dimensions[4].height = 18
  st.session_state.total_height += (18 * 4) + 36
  return worksheet

def headerBorders(worksheet):
  worksheet = box(worksheet, 'A1', 'B4')
  worksheet = box(worksheet, 'C1', 'C4')
  worksheet = box(worksheet, 'D1', 'F2')
  worksheet = box(worksheet, 'D3', 'F4')
  worksheet = box(worksheet, 'G1', 'G4')
  worksheet = box(worksheet, 'H1', 'H2')
  worksheet = box(worksheet, 'H3', 'H4')
  worksheet = box_fill(worksheet, 'I1', 'K4')
  return worksheet

def addHeaderValues(worksheet, projectName):
  worksheet.merge_cells('A1:B4')
  logo = padLogoImage('abi-logo.webp')
  worksheet.add_image(logo, 'A1')

  worksheet.merge_cells('E1:F2')
  worksheet['E1'] = f"{projectName.upper()}"
  worksheet['E1'].font = Font(name='Avenir Book', size=14, bold=True)
  worksheet['E1'].alignment = Alignment(horizontal='center', vertical='center')

  worksheet.merge_cells('E3:F4')
  worksheet['E3'] = "PLUMBING & BATH ACCESSORIES SCHEDULE"
  worksheet['E3'].font = Font(name='Avenir Book', size=12, bold=True)
  worksheet['E3'].alignment = Alignment(horizontal='center', vertical='center')

  worksheet.merge_cells('G2:G4')
  worksheet['G2'] = "S1100"
  worksheet['G2'].font = Font(name='Avenir Book', size=20, bold=True)
  worksheet['G2'].alignment = Alignment(horizontal='center', vertical='center')

  worksheet['C1'] = "63 Wingold Ave, Suite 208"
  worksheet['C1'].font = Font(name='Avenir Book', size=8)
  worksheet['C1'].alignment = Alignment(horizontal='center', vertical='center')

  worksheet['C2'] = "Toronto, ON, M6B 1P8"
  worksheet['C2'].font = Font(name='Avenir Book', size=8)
  worksheet['C2'].alignment = Alignment(horizontal='center', vertical='center')

  worksheet['C4'] = "alibuddinteriors.com"
  worksheet['C4'].font = Font(name='Avenir Book', size=8)
  worksheet['C4'].alignment = Alignment(horizontal='center', vertical='center')

  worksheet['D1'] = "PROJECT"
  worksheet['D1'].font = Font(name='Avenir Book', size=6)
  worksheet['D1'].alignment = Alignment(horizontal='left', vertical='center')

  worksheet['D3'] = "TITLE"
  worksheet['D3'].font = Font(name='Avenir Book', size=6)
  worksheet['D3'].alignment = Alignment(horizontal='left', vertical='center')

  worksheet['G1'] = "NO."
  worksheet['G1'].font = Font(name='Avenir Book', size=6)
  worksheet['G1'].alignment = Alignment(horizontal='left', vertical='center')

  worksheet['H1'] = "BY"
  worksheet['H1'].font = Font(name='Avenir Book', size=6)
  worksheet['H1'].alignment = Alignment(horizontal='left', vertical='center')

  worksheet['H2'] = "--"
  worksheet['H2'].font = Font(name='Avenir Book', size=8)
  worksheet['H2'].alignment = Alignment(horizontal='center', vertical='center')

  worksheet['H3'] = "DATE"
  worksheet['H3'].font = Font(name='Avenir Book', size=6)
  worksheet['H3'].alignment = Alignment(horizontal='left', vertical='center')

  worksheet['H4'] = "2024-06-14"
  worksheet['H4'].font = Font(name='Avenir Book', size=8)
  worksheet['H4'].alignment = Alignment(horizontal='center', vertical='center')

  worksheet['I1'] = "VERSION"
  worksheet['I1'].font = Font(name='Avenir Book', size=8)
  worksheet['I1'].alignment = Alignment(horizontal='left', vertical='center')

  worksheet['J1'] = "DATE"
  worksheet['J1'].font = Font(name='Avenir Book', size=8)
  worksheet['J1'].alignment = Alignment(horizontal='left', vertical='center')

  worksheet['K1'] = "DESCRIPTION"
  worksheet['K1'].font = Font(name='Avenir Book', size=8)
  worksheet['K1'].alignment = Alignment(horizontal='left', vertical='center')

  return worksheet

def mainTableHeaders(worksheet):
  worksheet = box_fill(worksheet, 'A6', 'K6')
  
  worksheet['A6'] = 'LOCATION'
  worksheet['A6'].font = Font(name='Avenir Book', size=8, bold=True)
  worksheet['A6'].alignment = Alignment(horizontal='center', vertical='center')
  worksheet['A6'].fill = fill

  worksheet['B6'] = 'ITEM #'
  worksheet['B6'].font = Font(name='Avenir Book', size=8, bold=True)
  worksheet['B6'].alignment = Alignment(horizontal='center', vertical='center')
  worksheet['B6'].fill = fill

  worksheet['C6'] = 'ITEM'
  worksheet['C6'].font = Font(name='Avenir Book', size=8, bold=True)
  worksheet['C6'].alignment = Alignment(horizontal='center', vertical='center')
  worksheet['C6'].fill = fill

  worksheet['D6'] = 'QTY'
  worksheet['D6'].font = Font(name='Avenir Book', size=8, bold=True)
  worksheet['D6'].alignment = Alignment(horizontal='center', vertical='center')
  worksheet['D6'].fill = fill

  worksheet['E6'] = 'IMAGE'
  worksheet['E6'].font = Font(name='Avenir Book', size=8, bold=True)
  worksheet['E6'].alignment = Alignment(horizontal='center', vertical='center')
  worksheet['E6'].fill = fill

  worksheet['F6'] = 'MODEL'
  worksheet['F6'].font = Font(name='Avenir Book', size=8, bold=True)
  worksheet['F6'].alignment = Alignment(horizontal='center', vertical='center')
  worksheet['F6'].fill = fill

  worksheet['G6'] = 'DIMENSIONS'
  worksheet['G6'].font = Font(name='Avenir Book', size=8, bold=True)
  worksheet['G6'].alignment = Alignment(horizontal='center', vertical='center')
  worksheet['G6'].fill = fill

  worksheet['H6'] = 'FINISH'
  worksheet['H6'].font = Font(name='Avenir Book', size=8, bold=True)
  worksheet['H6'].alignment = Alignment(horizontal='center', vertical='center')
  worksheet['H6'].fill = fill

  worksheet.merge_cells('I6:J6')
  worksheet['I6'] = 'SUPPLIER'
  worksheet['I6'].font = Font(name='Avenir Book', size=8, bold=True)
  worksheet['I6'].alignment = Alignment(horizontal='center', vertical='center')
  worksheet['I6'].fill = fill

  worksheet['K6'] = 'NOTES'
  worksheet['K6'].font = Font(name='Avenir Book', size=8, bold=True)
  worksheet['K6'].alignment = Alignment(horizontal='center', vertical='center')
  worksheet['K6'].fill = fill

  return worksheet

def createHeader(worksheet, projectName):
  worksheet = columnRowDimensions(worksheet)
  worksheet = headerBorders(worksheet)
  worksheet = addHeaderValues(worksheet, projectName)
  worksheet = mainTableHeaders(worksheet)
  return worksheet

@cache
def getSupplier(supplier):
  response = AI_client.chat.completions.create(
      model="gpt-3.5-turbo",
      messages=[
        {
          "role": "system",
          "content": (
            "You are an expert at extracting structured data from unstructured text. "
            "You always follow instructions exactly and never add extra commentary."
          )
        },
        {
          "role": "user",
          "content": f'''
  Extract the following details from the text below:
  - company name
  - salesperson
  - email
  - phone number

  Format your response as:
  company name>salesperson>email>phone number

  Rules:
  - Respond with ONLY the string in the format above, with exactly 3 '>' characters.
  - If any detail is missing, leave it blank but keep the separators (e.g. "company name>>email>").
  - Do NOT add any explanation, label, or extra text—just the string.

  Text:
  {supplier}
  '''
        }
      ],
      temperature=0.15
  )

  return response.choices[0].message.content

def addMainTable(worksheet, rooms, file, arranged_dataframe):
    st.session_state.pageStart = 0
    images = load_workbook(file).worksheets[0]._images[1:]
    total_rows = 0
    for room in rooms:
        for sub in rooms[room]:
            for _ in rooms[room][sub]:
                total_rows += 1
    images = pad_images_with_blanks(images, total_rows)

    # Tabloid (11x17), Landscape, Narrow margins
    AVAILABLE_HEIGHT_PER_PAGE = 800   # tweak as needed
    HEADER_HEIGHT = 108               # height of first-page header
    SUB_HEADER_HEIGHT = 18            # height of sub labels

    current_page_height = HEADER_HEIGHT
    row = 7
    pl = 1

    for main, subs in rooms.items():
        main_header_height = 18 

        if current_page_height + main_header_height > AVAILABLE_HEIGHT_PER_PAGE:
            worksheet.row_breaks.append(Break(id=row-1))
            current_page_height = HEADER_HEIGHT
            start = row
        worksheet.merge_cells(f'A{row}:K{row}')
        worksheet[f'A{row}'] = main
        worksheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        worksheet[f'A{row}'].font = Font(name='Avenir Book', size=8, bold=True)
        worksheet[f'A{row}'].fill = sec_fill
        row += 1
        current_page_height += main_header_height

        for sub in subs:
            start = row
            worksheet[f'A{row}'] = sub
            worksheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
            worksheet[f'A{row}'].font = Font(name='Avenir Book', size=8)

            has_products = False

            for product_type in subs[sub]:
                (type_product, index_product) = product_type
                row_df = arranged_dataframe[
                    arranged_dataframe['Area'].str.contains(main) &
                    arranged_dataframe['Area'].str.contains(sub)
                ]
                row_df = row_df[row_df['Image Index'] == index_product]

                if row_df.empty:
                    continue

                has_products = True

                brand = row_df.iloc[0].loc['Brand'] if pd.notna(row_df.iloc[0].loc['Brand']) else 'Not Provided'
                name = row_df.iloc[0].loc['Product Name'] if pd.notna(row_df.iloc[0].loc['Product Name']) else 'Not Provided'
                product_code = row_df.iloc[0].loc['Product Code #'] if pd.notna(row_df.iloc[0].loc['Product Code #']) else 'Not Provided'

                brand_lines = count_wrapped_lines('BRAND: ' + brand, MAX_MODEL_CHARACTERS_PER_COLUMN)
                name_lines = count_wrapped_lines('NAME: ' + name, MAX_MODEL_CHARACTERS_PER_COLUMN)
                sku_lines = count_wrapped_lines('SKU: ' + product_code, MAX_MODEL_CHARACTERS_PER_COLUMN)

                modelHeight = brand_lines + name_lines + sku_lines

                stringSample = CellRichText(
                    TextBlock(InlineFont(b=True), 'BRAND: '),
                    f'{brand}\n',
                    TextBlock(InlineFont(b=True), 'NAME: '),
                    f'{name}\n',
                    TextBlock(InlineFont(b=True), 'SKU: '),
                    f'{product_code}',
                )

                worksheet[f'F{row}'] = stringSample
                worksheet[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'F{row}'].font = Font(name='Avenir Book', size=8)

                worksheet[f'D{row}'] = row_df.iloc[0].loc['QTY (per Area)']
                worksheet[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                worksheet[f'D{row}'].font = Font(name='Avenir Book', size=8)

                worksheet[f'H{row}'] = row_df.iloc[0].loc['Finish/Color']
                worksheet[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'H{row}'].font = Font(name='Avenir Book', size=8)

                worksheet[f'G{row}'] = row_df.iloc[0].loc['Dimension']
                worksheet[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'G{row}'].font = Font(name='Avenir Book', size=8)

                worksheet.merge_cells(f'I{row}:J{row}')

                supplier_string = getSupplier(row_df.iloc[0].loc['Supplier'])
                details = [item.strip() for item in supplier_string.split('>') if item.strip()]

                supplierHeight = 0
                for detail in details:
                    supplierHeight += count_wrapped_lines(detail, MAX_SUPPLIER_CHARACTERS_PER_COLUMN)

                finalString = CellRichText(
                    TextBlock(InlineFont(b=True), f'{details[0]}\n'),
                    '\n'.join(details[1:])
                )

                worksheet[f'I{row}'] = finalString
                worksheet[f'I{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'I{row}'].font = Font(name='Avenir Book', size=8)

                worksheet[f'B{row}'] = f'PL-{pl}'
                worksheet[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'B{row}'].font = Font(name='Avenir Book', size=8)

                worksheet[f'K{row}'] = 'SEE ID DRAWINGS & SPEC SHEETS FOR MORE INFORMATION'
                worksheet[f'K{row}'].alignment = Alignment(horizontal='center', vertical='center')
                worksheet[f'K{row}'].font = Font(name='Avenir Book', size=8, italic=True)

                (worksheet[f'C{row}'], _) = product_type
                worksheet[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'C{row}'].font = Font(name='Avenir Book', size=8)

                max_lines = max(modelHeight, supplierHeight)
                row_height = max(36, min(300, max_lines * 18))

                if current_page_height + row_height > AVAILABLE_HEIGHT_PER_PAGE:
                    if start < row:
                        worksheet.merge_cells(f'A{start}:A{row-1}')
                    worksheet.row_breaks.append(Break(id=row-1))  # FIXED

                    current_page_height = SUB_HEADER_HEIGHT
                    start = row

                    worksheet[f'A{row}'] = sub
                    worksheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    worksheet[f'A{row}'].font = Font(name='Avenir Book', size=8)
                    worksheet.row_dimensions[row].height = SUB_HEADER_HEIGHT

                worksheet.row_dimensions[row].height = row_height
                st.session_state.total_height += row_height
                current_page_height += row_height

                image = images[row_df.iloc[0].loc["Image Index"]]
                if image.width != 1:
                    image = padProductImage(image, row_height)
                worksheet.add_image(image, f'E{row}')

                pl += 1
                row += 1

            if has_products and start < row:
                worksheet.merge_cells(f'A{start}:A{row-1}')

            main_keys = list(rooms.keys())
            sub_keys = list(subs.keys())
            is_last_sub_in_main = sub_keys.index(sub) == len(sub_keys) - 1
            if not is_last_sub_in_main:
                separator_height = 4.5
                if current_page_height + separator_height > AVAILABLE_HEIGHT_PER_PAGE:
                    worksheet.row_breaks.append(Break(id=row-1))  # FIXED
                    current_page_height = separator_height
                else:
                    current_page_height += separator_height

                worksheet.merge_cells(f'A{row}:K{row}')
                worksheet.row_dimensions[row].height = separator_height
                st.session_state.total_height += separator_height
                worksheet[f'A{row}'].fill = sec_fill
                row += 1

    worksheet = box_fill(worksheet, f'A7', f'K{row-1}')
    worksheet.freeze_panes = 'A7'
    return worksheet

@st.dialog("Log out")
def user_info(name):
  st.write(f'Hello, **{name}**!')
  if st.button("Log out", type="primary", use_container_width=True):
    st.logout()

if not st.user.is_logged_in:
  with st.container(horizontal_alignment="center"):
    with st.container(border=True, horizontal_alignment="center", width=500):
      st.image("abi-logo.webp")
      st.title("Export to Schedule App", width="content", anchor=False)
      st.text("Format to your needs.", width="content")
      st.write('---')
      login_btn = st.button(
        "🔑 **Log in** with **Google**",
        use_container_width=True,
        type="primary"
      )
      st.caption("Please Login to Continue", width="content")
      if login_btn:
        st.login()
else:
  user_details = st.user.to_dict()

  with st.container(border=False, horizontal=True, horizontal_alignment="distribute", height=150):
    with st.container(vertical_alignment="center", height="stretch"):
      st.image("abi-logo.webp", width=250)

    with st.container(vertical_alignment="center", height="stretch", horizontal_alignment="right"):
      logout_btn = st.button(
        f"☰",
        type="primary"
      )
      
      if logout_btn:
        user_info(user_details.get("name"))
  with st.container(border=True):
    st.header("Export to Formatted Schedule Application", anchor=False)
    with st.expander("Upload Schedule", expanded=True):
      st.session_state.input_file = st.file_uploader("Please input the raw export file", type=["xlsx", "xls"])
    if st.session_state.input_file:
      [projectName, ext] = st.session_state.input_file.name.split('-', 1)
      projectName = projectName.strip()
      projectCategory = ext.split('Fohlio')[0].strip()
      st.session_state.details = pd.read_excel(st.session_state.input_file, header=9)
      st.session_state.rooms = {}
      for idx, roomName in enumerate(st.session_state.details['Area']):
        [main, sub] = roomName.split(' / ')
        main_key = main[4:]
        sub_key = sub[4:]
        product = st.session_state.details['Product Type'][idx]
        product_name = st.session_state.details['Product Name'][idx]
        if main_key not in st.session_state.rooms:
          st.session_state.rooms[main_key] = {sub_key: [(product, idx)]}
        else:
          if sub_key not in st.session_state.rooms[main_key]:
            st.session_state.rooms[main_key][sub_key] = [(product, idx)]
          else:
            st.session_state.rooms[main_key][sub_key].append((product, idx))
      st.session_state.room_order = list(st.session_state.rooms.keys())

      column1, column2, column3 = st.columns(3)
      with column1:

        st.subheader("Floor Order", anchor=False)
        st.caption("Drag and rearrange the rooms as needed for your export. The order will be reflected in the final formatted schedule.")
        new_order = sortables.sort_items(
          st.session_state.room_order,
          direction="vertical"
        )

        if new_order != st.session_state.room_order:
          st.session_state.room_order = new_order
      with column2:
        st.subheader("Room Order", anchor=False)
        st.session_state.selectedRoom = st.selectbox(
          "Select a floor to view its rooms",
          options=st.session_state.room_order,
          index=0
        )
        if 'sub_room_orders' not in st.session_state:
          st.session_state.sub_room_orders = {}
        
        current_sub_rooms = list(st.session_state.rooms[st.session_state.selectedRoom].keys())
        
        if st.session_state.selectedRoom not in st.session_state.sub_room_orders:
          st.session_state.sub_room_orders[st.session_state.selectedRoom] = current_sub_rooms
        
        st.session_state.rooms_in_floor = st.session_state.sub_room_orders[st.session_state.selectedRoom]
        
        new_sub_order = sortables.sort_items(
          st.session_state.rooms_in_floor,
          direction="vertical"
        )
        
        if new_sub_order != st.session_state.rooms_in_floor:
          st.session_state.sub_room_orders[st.session_state.selectedRoom] = new_sub_order
          st.session_state.rooms_in_floor = new_sub_order

      with column3:
        st.subheader(f"File Input Details", anchor=False)
        st.write(f"**Project Name:** {projectName}")
        st.write(f"**Project Category:** {projectCategory}")
        st.info("Note that the application detects the **Project Name** and **Project Category** by the filename _`<project_name> - <project_category> Fohlio Raw Export.xlsx`_", icon="ℹ️")
      
      st.write('---')
      st.session_state.details['Image Index'] = range(len(st.session_state.details))
      st.session_state.details['main_area'] = st.session_state.details['Area'].apply(lambda x: x.split(' / ')[0][4:])
      room_order_map = {room: i for i, room in enumerate(st.session_state.room_order)}
      st.session_state.details['main_order'] = st.session_state.details['main_area'].map(room_order_map)
      st.session_state.details = st.session_state.details.sort_values('main_order').reset_index(drop=True)
      st.session_state.details = st.session_state.details.drop(columns=['main_area', 'main_order'])
      st.subheader("File Input Preview", anchor=False)
      st.write("You can preview the data extracted from your uploaded file below. This data will be used to generate the formatted schedule.")
      st.write(st.session_state.details)

      ordered_rooms = OrderedDict()
      for key in st.session_state.room_order:
        if key in st.session_state.rooms:
          ordered_rooms[key] = st.session_state.rooms[key]
      for sub_key in new_sub_order:
        if sub_key in ordered_rooms[st.session_state.selectedRoom]:
          ordered_rooms[st.session_state.selectedRoom][sub_key] = ordered_rooms[st.session_state.selectedRoom].pop(sub_key)

      newDataframe = pd.DataFrame()
      output = BytesIO()
      with st.spinner("Please wait while we format your schedule..."):
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
          st.session_state.total_height = 0
          newDataframe.to_excel(writer, index=False, sheet_name=projectCategory)
          worksheet = writer.sheets[projectCategory]
          worksheet = createHeader(worksheet, projectName)
          worksheet = addMainTable(worksheet, ordered_rooms, st.session_state.input_file, st.session_state.details)
          worksheet.page_setup.fitToPage = True
          worksheet.page_setup.fitToWidth = 1
          worksheet.page_setup.fitToHeight = False
          worksheet.page_margins.left = 0.25
          worksheet.page_margins.right = 0.25
          worksheet.page_margins.top = 0.75
          worksheet.page_margins.bottom = 0.75
          worksheet.page_setup.orientation = 'landscape'
          worksheet.page_setup.paperSize = 3
        output.seek(0)
        print(st.session_state.total_height)
        st.session_state.output = output
        st.download_button(
          label="Download Formatted Excel File",
          data=output,
          use_container_width=True,
          file_name=f"{projectName} - {projectCategory} Automated Export.xlsx",
          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
          type="primary"
        )
    else:
      st.info("Please upload a schedule file to proceed.")

st.markdown("""
<div style="text-align: center; color: #6c757d; font-size: 14px; margin-top: 30px;">
  <p style="margin-bottom: 0px;"><strong>Export to Formatted Schedule Application</strong></p>
  <a style="margin-bottom: 0px; text-decoration: none;" href="https://lernmoreconsulting.com">© 2025 LernMore Consulting</a>
  <p>Secure • Fast • Accurate</p>
</div>
""", unsafe_allow_html=True)